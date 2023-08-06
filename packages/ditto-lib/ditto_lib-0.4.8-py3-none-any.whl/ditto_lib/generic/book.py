#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  book.py
#  Developed in 2019 by Hernan Gelaf-Romer <hernan_gelafromer@student.uml.edu>
#
#  Implements the a container to store load excel workbooks.
#

try:
    import openpyxl as pyxl
except:
    pyxl = None

from ordered_set import OrderedSet
from ditto_lib.generic.utils import check_package
from ditto_lib.generic.itemcollection import ItemCollection, Attribute
from ditto_lib.generic.config import logger

class Book:

    '''
    Support for importing multi sheet, excel files into
    ItemCollections
    '''

    class FillDiff:

        '''
        Diff class for fill method
        '''

        def __init__(self, name, sheet, item, attribute, color, pattern_type, fill_type):
            self.name = name
            self.sheet = sheet
            self.item = item
            self.attribute = attribute
            self.color = color
            self.pattern_type = pattern_type
            self.fill_type = fill_type

        def apply(self, cell):
            logger.log('debug', "Applying diff: {}".format(self.name))
            cell.fill = pyxl.styles.PatternFill(
                start_color=self.color[0],
                end_color=self.color[1],
                fill_type=self.fill_type,
                patternType=self.pattern_type
            )

    def __init__(self, name):
        check_package(pyxl, 'openpyxl', 'Book Module')
        self._sheets = {}
        self._diff_tracker = {}
        self._name = name
        self._offsets = []

    @property
    def name(self):
        '''
        Name of the book
        '''
        return self._name

    @name.setter
    def name(self, name):
        '''
        Set name of the Book
        '''
        logger.log('debug', "{} name set to {}".format(self._name, name))
        self._name = name

    def collection(self, name):
        '''
        Retrieve the item collection associated 
        with the given name
        '''
        if name in self._sheets:
            return self._sheets[name]
        else:
            logger.log('error', "No sheetname called {} in {}".format(name, self.name))
            raise ValueError("No sheetname called {} in {}".format(name, self.name))

    @property
    def sheetnames(self):
        '''
        Return a list of all the sheet names associated
        with this Book
        '''
        return list(self._sheets.keys())
        
    def load(self, filename, start_indexes, non_descriptors=set()):
        '''
        Load the given excel file. Assumes all item names are located in the first column\n
        Args:\n
        filename: The excel file to load\n
        start_indexes: A list of tuples, where the first item in the tuple is the start row index
        of the current sheet, and the second item is the start column index of the current sheet\n
        non_descriptors: The set of attributes which will be defined as non_descriptors
        '''
        self.wipe()
        self._offsets = start_indexes
        if '.xlsx' not in filename:
            filename += '.xlsx'
        wb = pyxl.load_workbook(filename, read_only=True)
        logger.log('debug', "Sucessfully loaded excel file {}, with sheets {}".format(filename, wb.sheetnames))
        for indexes, sheet in zip(start_indexes, wb.sheetnames):
            start_row = indexes[0]
            start_column = indexes[1]
            collection = ItemCollection(sheet)
            for cell in wb[sheet][start_row][start_column:]:
                collection._attributes.add(Attribute(cell.value, cell.value not in non_descriptors))
            logger.log('debug', "Attributes {} generated for {}".format(
                [collection._attributes[i].name for i in range(len(collection._attributes))],
                sheet))
            for row in wb[sheet].iter_rows(min_row=start_row + 1):
                item_name = row[0].value
                collection.items[item_name] = []
                for cell in row[start_column:]:
                    collection.items[item_name].append(cell.value)
            self._sheets[sheet] = collection
        logger.log('info', "Done loading {}".format(filename))

    def save(self, filename):
        '''
        Save the book to an excel file with the given
        file name
        '''
        if '.xlsx' not in filename:
            filename += '.xlsx'
        logger.log('debug', "Creating save WorkBook")
        wb = pyxl.Workbook(write_only=True)

        for sheetname, collection in self._sheets.items():
            ws = wb.create_sheet(sheetname)
            ws.append(['Name'] + [collection._attributes[i].name for i in range(len(collection._attributes))])

            if sheetname not in self._diff_tracker:
                for item, values in collection._items.items():
                    ws.append([item] + values)
            else:
                for item, values in collection._items.items():
                    row_values = []
                    for idx, value in enumerate(values):
                        attribute = collection._attributes[idx]
                        diff_loc = item + attribute.name
                        if diff_loc not in self._diff_tracker[sheetname]:
                            row_values.append(value)
                        else:
                            cell = pyxl.cell.WriteOnlyCell(ws, value=value)
                            for diff in self._diff_tracker[sheetname][diff_loc]:
                                diff.apply(cell)
                            row_values.append(cell)
                            del self._diff_tracker[sheetname][diff_loc]
                    ws.append([item] + row_values)
                
        logger.log('debug', "Save WorkBook created")
        logger.log('info', "Saving Book {}".format(filename))
        wb.save(filename=filename)
        logger.log('info', "Book saved to {}".format(filename))

    def add_sheet(self, itemcollection):
        '''
        Add a sheet to the Book, takes an 
        item collection
        '''
        self._sheets[itemcollection.name] = itemcollection

    def value(self, sheet, item, attribute):
        '''
        Retrieve the value\n
        Args:\n
        sheet: Name of the sheet\n
        item: Name of the item\n
        attribute: Name of the attribute
        '''
        return self.collection(sheet).item_attribute(item, attribute)

    def set_value(self, sheet, item, value, attribute_name, is_descriptor=False):
        '''
        Set the value\n
        Args:\n
        sheet: Name of the sheet\n
        item: Name of the item\n
        attribute_name: Name of the attribute\n
        value: Attribute value\n
        is_descriptor: Boolean whether attribute is a descriptor
        '''
        self.collection(sheet).set_item_attribute(item, value, attribute_name, is_descriptor)

    def merge(self, sheets=None):
        '''
        Return an ItemCollection that is the results of 
        merging all the sheets in the given sheets arg. Defaults
        to None which will return all the sheets merged
        '''
        if sheets is None:
            sheets = self.sheetnames

        collection = self.collection(sheets[0])
        for sheet in sheets[1:]:
            collection = collection.merge(self.collection(sheet), 'Merged Collection')
        return collection

    def intersect(self, sheets=None):
        '''
        Return an ItemCollection that is the results of 
        intersecting all the sheets in the given sheets arg. Defaults
        to None which will return all the sheets intersected
        '''
        if sheets is None:
            sheets = self.sheetnames

        collection = self.collection(sheets[0])
        for sheet in sheets[1:]:
            collection = collection.intersect(self.collection(sheet), 'Intersected Collection')
        return collection

    def fill(self, sheet, item, attribute, color, pattern_type='solid', fill_type='solid'):
        '''
        Fill a cell in the excel sheet with a color\n
        Args:\n
        sheet: The sheet which stores cell\n
        item: The item name\n
        attribute: The attribute name\n
        color: A tuple, the first item is the start_color, the second item is 
        the end_color. More information on styling can be found at openpyxl's
        website\n
        pattern_type: The type of pattern the fill will hold. More options can be 
        found at openpyxl's website. Defaults to 'solid'\n
        fill_type: The fill_type. More information can be found at openpyxl's website.
        Defaults to 'solid'
        '''
        if sheet not in self._diff_tracker:
            self._diff_tracker[sheet] = {}
        if (item + attribute) not in self._diff_tracker[sheet]:
            self._diff_tracker[sheet][item + attribute] = []
        self._diff_tracker[sheet][item + attribute].append(self.FillDiff(
            name="'Fill sheet: {}, item: {}, attribute: {}'".format(sheet, item, attribute),
            sheet=sheet,
            item=item, 
            attribute=attribute,
            color=color,
            pattern_type=pattern_type,
            fill_type=fill_type
        ))

    def wipe(self):
        '''
        Reset this book
        '''
        self._sheets = {}
        self._diff_tracker = {}
        logger.log('debug', "{} reset".format(self.name))
        