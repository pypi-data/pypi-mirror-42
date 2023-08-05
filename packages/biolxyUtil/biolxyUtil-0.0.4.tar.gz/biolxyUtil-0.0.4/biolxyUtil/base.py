#!/usr/bin/env python
# -*- coding: utf-8 -*-
u"""
File Name   : biolxyUtil.py .

Author      : biolxy
E-mail      : biolxy@aliyun.com
Created Time: 2018-07-24 13:34:17
version     : 1.0
Function    : 用来提供一些基础功能，例如linux客户端彩色输出，记录日志，安全shell命令等
"""
import shlex
import logging
import subprocess
from collections import Counter
from Bio import SeqIO
from Bio.Seq import Seq
import time
import sys
import os
import re
from multiprocessing import cpu_count


class MagicDict(dict):
    u"""
    Mdict = MagicDict()
    Mdict[a][b] = c

    Mdict = MagicDict(dict1)  # 转化一个dict为MagicDict
    """

    def __getitem__(self, item):
        try:
            return dict.__getitem__(self, item)
        except KeyError:
            value = self[item] = type(self)()
            return value

class Timer:
    # u"""计时器.
    # 仅仅适用python3
    # 参见：
    # https://python3-cookbook.readthedocs.io/zh_CN/latest/c13/p13_making_stopwatch_timer.html
    # python3 的话可以让 func=time.perf_counter ，能获得更高的时间精度
    # """
    def __init__(self, func=time.time):
        self.elapsed = 0.0
        self._func = func
        self._start = None

    def start(self):
        if self._start is not None:
            raise RuntimeError('Already started')
        self._start = self._func()

    def stop(self):
        if self._start is None:
            raise RuntimeError('Not started')
        end = self._func()
        self.elapsed += end - self._start
        self._start = None

    def reset(self):
        self.elapsed = 0.0

    @property
    def running(self):
        return self._start is not None

    def __enter__(self):
        self.start()
        return self

    def __exit__(self, *args):
        self.stop()


class progressPrintStr:
    def __init__(self, inputStr, func=time.time, func2=time.localtime, func3=time.strftime):
        self.elapsed = 0.0
        self._func = func
        self._func2 = func2
        self._func3 = func3
        self._start = None
        self.inputStr = inputStr

    def start(self):
        if self._start is not None:
            raise RuntimeError('Already started')
        self._start = self._func()
        self._startstr1 = self._func3('%Y-%m-%d %H:%M:%S', self._func2(self._start))
        self._startstr2 = "   [ {} is begin at time: {} ]   ".format(self.inputStr, self._startstr1)
        self._startline = len(self._startstr2) * "="
        print(color_term(self._startline, "green", True))
        print(color_term(self._startstr2, "green", True))
        print(color_term(self._startline, "green", True))


    def stop(self):
        if self._start is None:
            raise RuntimeError('Not started')
        self._end = self._func()
        self.elapsed += self._end - self._start
        self.elapsedMinutes = int(self.elapsed / 60 )
        self.elapsedSeconds = self.elapsed % 60
        self._start = None
        self._endstr1 = self._func3('%Y-%m-%d %H:%M:%S', self._func2(self._end))
        self._endstr2 = "   [ {} is end at time: {}, {} min {:.2f} s has elapsed ]   ".format(
            self.inputStr,
            self._endstr1,
            self.elapsedMinutes,
            self.elapsedSeconds)
        self._endline = len(self._endstr2) * "="
        print(color_term(self._endline, "green", True))
        print(color_term(self._endstr2, "green", True))
        print(color_term(self._endline, "green", True))

    def reset(self):
        self.elapsed = 0.0

    @property
    def running(self):
        return self._start is not None

    def __enter__(self):
        self.start()
        return self

    def __exit__(self, *args):
        self.stop()


class StreamToLogger(object):
    u"""
    Fake file-like stream object that reinputDirects writes to a logger instance.

    记录log日志，适配py3
    Refer: [ReinputDirect stdout and stderr to a logger in Python]
           (https://www.electricmonk.nl/log/2011/08/14/reinputDirect-stdout-and-stderr-to-a-logger-in-python/)
           [How to reinputDirect stdout and stderr to logger in Python]
           (https://stackoverflow.com/questions/19425736/how-to-reinputDirect-stdout-and-stderr-to-logger-in-python)
    """

    def __init__(self, logger, log_level=logging.INFO):
        self.logger = logger
        self.log_level = log_level
        self.linebuf = ''

    def write(self, buf):
        for line in buf.rstrip().splitlines():
            self.logger.log(self.log_level, line.rstrip())
    # logging.basicConfig(level=logging.DEBUG,format='%(asctime)s:%(levelname)s:%(name)s:%(message)s',filename="out.log",filemode='a')
    # SetupLogger('log', os.path.join(inputDir_output, 'handleAutobox.log'), log_mode=args.log_mode,
    #             format_sh='%(message)s', format_fh='[%(levelname)s] %(message)s')
    # logger = logging.getLogger('log')


def getcpuNumber():
    number = cpu_count()
    return number


def color_term(string, color='blue', bold=True):
    u"""Linux客户端彩色输出，适配py2，py3."""
    colors = {
        'grey': '\033[0;30m',
        'red': '\033[0;31m',
        'green': '\033[0;32m',
        'yellow': '\033[0;33m',
        'blue': '\033[0;34m',
        'megenta': '\033[0;35m',
        'cyan': '\033[0;36m',
        'white': '\033[0;37m',
        'bold': '\033[1m',
        'end': '\033[0m'
    }
    color_format = colors[color].replace('[0', '[1') if bold else colors[color]
    return color_format + str(string) + colors['end']
    # print(color_term("{}".format(var), 'red'))





def execute_cmd(cmd, printStstus=True):
    u"""Change sys.system(),提供安全的shell输入端口，为以后web键入命令提供基础,适配py2, py3.

    execute_cmd 中可以直接嵌套 linux命令，同样可以嵌套类似 python script.py inputfile 等命令
    通常用法为 execute_cmd("mkinputDir {}".format())
    该函数调用 color_term 函数
    """
    # https://python3-cookbook.readthedocs.io/zh_CN/latest/c13/p06_executing_external_command_and_get_its_output.html
    start_time = time.time()
    start_time2 = time.strftime(
        '[ %Y-%m-%d %H:%M:%S ]', time.localtime(start_time))
    print(color_term("{} Command will be execute in a subshell:\n{}".format(
        start_time2, cmd), 'cyan', bold=False))
    returnNum = 1
    try:
        p = subprocess.Popen(shlex.split(cmd),
                             stdout=subprocess.PIPE,
                             stderr=subprocess.PIPE,
                             shell=False)
        # To interpret as text, decode
        if printStstus == True:
            for line in iter(p.stdout.readline, b''):
                print(line.rstrip().decode())
                sys.stdout.flush()
        else:
            pass
        for line in iter(p.stderr.readline, b''):
            sys.stderr.flush()
            print(color_term("{}".format(line.rstrip().decode()), 'yellow', bold=False))
        p.wait()
        returnNum = int(p.returncode)
        if returnNum != 0:
            print(color_term("This command returnNum is {}".format(returnNum), 'red'))
    except BaseException as e:
        returnStstus = False
        raise BaseException(color_term(e, 'red'))
    return returnNum


def execute_cmd2(cmd):
    returnNum = 1
    start_time = time.time()
    start_time2 = time.strftime(
        '[ %Y-%m-%d %H:%M:%S ]', time.localtime(start_time))
    print(color_term("{} Command will be execute in a subshell:\n{}".format(
        start_time2, cmd), 'cyan', bold=False))
    returnStstus = True
    try:
        returnNum = os.system(cmd)
    except Exception as e:
        returnStstus = False
        raise Exception(color_term(e, 'red'))
    if returnNum != 1:
        returnNum = returnNum >> 8
    return returnNum

def diff_multiple_list(list_list):
    u"""
    对多个数组求交集，并集
    """
    listall = []
    for listi in list_list:
        listall += listi
    # Intersection
    list_A = []
    dictlist = Counter(listall)
    numof_list = len(list_list)
    for i in dictlist:
        if dictlist[i] == numof_list:
            list_A.append(i)
    # Union
    list_B = list(set(listall))
    return list_A, list_B





def getDictbyListFromFasta(inputFasta, inputList):
    # def readFasta2dict(infile):
    #     sequence = []
    #     fastaDict = {}
    #     with open(infile) as file_one:
    #         file_one_content = file_one.read()
    #         for line in file_one_content.split("\n"):
    #             if not line.strip():
    #                 continue
    #             if line.startswith(">"):
    #                 sequence_name = line.rstrip('\n')[1:]
    #             else:
    #                 sequence.append(line.rstrip('\n'))
    #             fastaDict[sequence_name] = "".join(sequence)
    #     return fastaDict
    # 内存消耗过于巨大
    sequenceName = []
    fastaDict = {}
    inputFastaDict = SeqIO.index(inputFasta, "fasta")
    for i in inputList:
        SeqRecord = inputFastaDict[i]
        fastaDict[i] = str(SeqRecord.seq)
    return fastaDict


def getTranslatePep(inputDict):
    def getTP(sequence):
        inputStr = ""
        seq = Seq(sequence)
        proseq = seq.translate()
        startNumber = proseq.find("M")
        endNumber = proseq.find("*", startNumber, len(proseq) + 1)
        strPro = str(proseq)
        # print(strPro[startNumber:endNumber])
        return strPro[startNumber:endNumber]

    returnDict = {}
    for item in inputDict:
        seq1 = inputDict[item]
        seq2 = seq1[1:]
        seq3 = seq1[2:]
        dict1 = {}
        list1Length = []
        for i in seq1, seq2, seq3:
            por = getTP(seq1)
            len1 = len(por)
            list1Length.append(len1)
            dict1[len1] = por
        if len(dict1[max(list1Length)]) >= 8:
            returnDict[item] = dict1[max(list1Length)]
    return returnDict


def get_real_path(pathfile):
    if os.path.islink(pathfile):
        pathfile = os.readlink(pathfile)
    pathfile = os.path.abspath(pathfile)
    return pathfile


def create_bam_index(pathbam):
    pathbambai = pathbam + ".bai"
    if not os.path.exists(pathbambai):
        try:
            cmd = "sambamba index {}".format(pathbam)
            execute_cmd
        except Exception as e:
            raise Exception(color_term(e, 'red'))



def getFile(inputDir, inputStr):
    list1 = []
    if os.path.isdir(inputDir):
        for relpath, dirs, files in os.walk(inputDir):
            for item in files:
                if re.search(inputStr, item):
                    full_path = os.path.join(inputDir, relpath, item)
                    str1 = os.path.normpath(os.path.abspath(full_path))
                    print(str1)
                    list1.append(str1)
    else:
        print("{} is not a dir".format(inputDir))
    return list1

def progressBar(number, time1):
    time1 = float(time1)
    number = int(number)
    for i in range(number):
        # 进度条类型
        sys.stdout.write("*")
        sys.stdout.flush()
        time.sleep(time1)

def progressBar2(number, time1):
    time1 = float(time1)
    number = int(number) + 1
    for i in range(number):
        sys.stdout.write('   \r')
        sys.stdout.flush()
        sys.stdout.write('{}%\r'.format(i))
        sys.stdout.flush()
        time.sleep(time1)



def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    From: https://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data-using-pandas
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    attention:
        对于Pandas <0.21.0，请替换sheet_name为sheetname！
    Expample:
        append_df_to_excel('d:/temp/test.xlsx', df)
        append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)
        append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', index=False)
        append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', index=False, startrow=25)
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def getColumnNamesNum(infile):
    """
    input infile, get a dict (key is cloumnName , value is the cloumn number )
    """
    linelist = []
    dict1 = {}
    with open(infile) as ff:
        for line in ff:
            line = line.rstrip()
            linelist = line.split("\t")
            break
    for x, columnName in enumerate(linelist):
        dict1[columnName] = x
    return dict1