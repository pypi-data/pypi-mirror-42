#!/usr/bin/env python3
"""Dump data to xlsx file.

.dump_drill_pass()
.dump_ktu()
.dump_salary()
"""

from __future__ import annotations

import time
import pandas as pd
from pathlib import PurePath
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from .standart_functions import BasicFunctionsS as BasF_S


class DumpToExl(BasF_S):
    """Dump data to xlsx file."""

    __slots__ = ()

    months = [
        '', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
        'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря'
    ]

    @classmethod
    def _create_pass_name(cls, passport) -> str:
        """Create passport name."""
        pass_name = (
            "{}-{}-{} {}"
            .format(
                passport.params.year,
                passport.params.month,
                passport.params.day,
                int(passport.params.number)
            )
        )
        return pass_name

    @classmethod
    def _check_dir(cls, dir: PurePath):
        """Create empty directory if not exists."""
        if not dir.exists():
            dir.mkdir(parents=True)

    @classmethod
    def _normilise_barehole(cls, bareholes):
        """Normalise 5+ meters bareholes to 5 meters."""
        count = 0
        new_bareholes = {}
        for key in bareholes:
            if key >= 5:
                count += bareholes[key]
            else:
                new_bareholes[key] = bareholes[key]
        if count:
            new_bareholes[5] = count
        return new_bareholes

    @classmethod
    def _dump_real_salary(cls, report: 'MainReport'):
        """Dumpp real salary to exel blanc."""
        real_salary_path = (
            super().get_root_path().parent / 'Documents' / 'Табеля'
        )
        cls._check_dir(real_salary_path)
        real_salary_blanc_path = (
            super().get_root_path() / 'exl_blancs' / 'real_salary.xlsx'
        )
        workbook = load_workbook(real_salary_blanc_path)
        worksheet = workbook.active
        worksheet['C1'] = report.status['date']

        salary = report.workers_showing['факт']['зарплата']
        for worker_number, worker in enumerate(salary, start=1):
            row_number = 4 + worker_number
            worksheet['B' + str(row_number)] = worker_number
            worksheet['C' + str(row_number)] = super().make_name_short(worker)
            worksheet['D' + str(row_number)] = salary[worker]

        name = f"{report.status['date']} {report.status['shift']} на руки"
        pass_name = real_salary_path.joinpath(name).with_suffix('.xlsx')
        workbook.save(pass_name)

    @classmethod
    def _fill_salary(cls, workbook, report, user):
        """Fill salary to exel."""
        worksheet = workbook.active
        brigadiers_path = super().get_root_path() / 'data' / 'brigadiers'
        brigadiers = super().load_data(
            data_path=brigadiers_path,
            user=user,
        )
        worksheet['G28'] = report.totall
        ktu = report.workers_showing['бух.']['КТУ']
        hours = report.workers_showing['бух.']['часы']
        salary = report.workers_showing['бух.']['зарплата']

        for worker_number, worker in enumerate(ktu, start=1):
            row_number = 30 + worker_number
            worksheet['B' + str(row_number)] = worker_number
            worksheet['C' + str(row_number)] = super().make_name_short(worker)
            worksheet['D' + str(row_number)] = hours[worker]
            worksheet['G' + str(row_number)] = ktu[worker]
            addition = 0
            if worker in brigadiers:
                addition = 0.15
            worksheet['M' + str(row_number)] = addition
            worksheet['P' + str(row_number)] = salary[worker]
            worksheet['J' + str(row_number)] = (
                salary[worker] / (addition+1)
            )

    @classmethod
    def dump_drill_pass(cls, passport, negab=None):
        """Dump drill passport data to blanc exl file."""
        blanc_drill_path = (
            super().get_root_path() / 'exl_blancs' / 'drill_passport.xlsx')
        drill_pass_path = (
            super().get_root_path().parent / 'Documents' / 'Буровые_паспорта')
        cls._check_dir(drill_pass_path)
        workbook = load_workbook(blanc_drill_path)
        worksheet = workbook.active
        if negab:
            img = Image(
                super().get_root_path() / 'exl_blancs' / 'scheme_ng.png'
            )
            worksheet['F4'] = 'колличество негабаритов:'
            worksheet['K4'] = int(negab)
        else:
            img = Image(super().get_root_path() / 'exl_blancs' / 'scheme.png')

        worksheet.add_image(img, 'A29')
        worksheet['K1'] = int(passport.params.number)  # Passport number.
        worksheet['J5'] = str(passport.params.day)  # Day.
        worksheet['K5'] = cls.months[int(passport.params.month)]  # Month.
        worksheet['M5'] = str(passport.params.year)  # Year.
        worksheet['Q6'] = str(passport.params.horizond)  # Horizond.
        worksheet['F9'] = float(passport.params.pownder)  # Pownder.
        worksheet['K9'] = int(passport.params.d_sh)  # D_SH.
        worksheet['P9'] = int(passport.params.detonators)  # Detonators.
        # Bareholes.
        norm_bareholes = cls._normilise_barehole(passport.bareholes)
        for row_number, length in enumerate(norm_bareholes, start=1):
            row_number += 15
            worksheet['G' + str(row_number)] = length
            worksheet['D' + str(row_number)] = int(norm_bareholes[length])
        # Volume
        volume = round(
            float(passport.params.pownder) * 5
            + (float(passport.params.d_sh) / 10), 1
        )
        worksheet['K27'] = volume
        # Block params.
        height = float(passport.params.block_height)
        worksheet['H25'] = height
        if float(passport.params.block_depth) > 5:
            depth = 5
        else:
            depth = float(passport.params.block_depth)
        worksheet['P25'] = depth
        worksheet['L25'] = round(volume / height / depth, 1)
        worksheet['M8'] = (
            round(
                (worksheet['L25'].value - 0.4)
                / ((passport.params.block_width - 0.4) / 0.35), 3
            ) * 1000
        )
        # Master.
        master = super().make_name_short(str(passport.params.master))
        worksheet['J47'] = master
        # Save file.
        pass_name = cls._create_pass_name(passport)
        pass_name = drill_pass_path.joinpath(pass_name).with_suffix('.xlsx')
        workbook.save(pass_name)
        print("\nФайл сохранен:\n", str(pass_name))
        time.sleep(3)

    @classmethod
    def dump_ktu(cls, report):
        """Dump KTU data to blanc exl file."""
        ktu_path = super().get_root_path().parent / 'Documents' / 'КТУ'
        cls._check_dir(ktu_path)
        blanc_ktu_path = super().get_root_path() / 'exl_blancs' / 'ktu.xlsx'
        ktu = report.workers_showing['бух.']['КТУ']
        hours = report.workers_showing['бух.']['часы']
        year = report.status['date'].split('-')[0]
        month = (
            cls.months[int(report.status['date'].split('-')[1][:2])][:-1]+'е'
        )
        shift = report.status['shift']
        brig_list = {
            'Смена 1': 'Бригадой №1',
            'Смена 2': 'Бригадой №2'
        }
        brig = brig_list[shift]

        workbook = load_workbook(blanc_ktu_path)
        worksheet = workbook.active
        worksheet['C4'] = brig
        worksheet['C5'] = month
        worksheet['D5'] = year
        for worker_number, worker in enumerate(ktu, start=1):
            row_number = 7 + worker_number
            worksheet['A' + str(row_number)] = worker_number
            worksheet['B' + str(row_number)] = super().make_name_short(worker)
            worksheet['C' + str(row_number)] = hours[worker]
            worksheet['D' + str(row_number)] = ktu[worker]
        # Save file.
        pass_name = '-'.join([
            year, report.status['date'].split('-')[1][:2], shift])
        pass_name = ktu_path.joinpath(pass_name).with_suffix('.xlsx')
        workbook.save(pass_name)
        print("\nФайл сохранен:\n", str(pass_name))
        time.sleep(3)

    @classmethod
    def dump_salary(cls, report: 'MainReport', user):
        """Dump Salary to exists exel tabel."""
        salary_path = super().get_root_path().parent / 'Documents' / 'Табеля'
        cls._check_dir(salary_path)
        name = report.status['date'] + ' ' + report.status['shift']
        find = None
        for file in salary_path.iterdir():
            if name in str(file):
                find = name
                break
        if find:
            file_path = salary_path.joinpath(find).with_suffix('.xlsx')
            workbook = load_workbook(file_path)
            cls._fill_salary(workbook, report, user)
            workbook.save(file_path)
            cls._dump_real_salary(report)
            print("\nФайл сохранен:\n", salary_path.joinpath(find))
        else:
            print("Табель не найден в папке 'Табеля'")
        time.sleep(3)

    @classmethod
    def dump_worker_salary(cls, salary: pd.DataFrame):
        """Dump non brigade workers salary to exel."""
        worker_salary_path = (
            super().get_root_path().parent / 'Documents' / 'Табеля'
        )
        cls._check_dir(worker_salary_path)
        worker_salary_blanc_path = (
            super().get_root_path() / 'exl_blancs' / 'workers_salary.xlsx'
        )
        workbook = load_workbook(worker_salary_blanc_path)
        worksheet = workbook.active

        worksheet['C1'] = cls.months[int(list(set(salary.month))[0])]

        for row_number, worker in enumerate(salary.name, start=1):
            row = 5 + row_number
            worker_name = super().make_name_short(worker)
            worker_salary = salary[salary.name == worker]
            worksheet['B' + str(row)] = row_number
            worksheet['C' + str(row)] = worker_name
            worksheet['D' + str(row)] = float(worker_salary['salary'])

        pass_name = worker_salary_path.joinpath(
            f"Суммы {worksheet['C1'].value} {list(set(salary['shift']))[0]}"
        ).with_suffix('.xlsx')
        workbook.save(pass_name)
        print("\nФайл сохранен:\n", str(pass_name))
        time.sleep(3)
